"""Workbook creation logic for Excel Dashboard Studio."""

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config.palettes import PALETTES
from core.logo_handler import add_logo_or_placeholder
from marketing.marketplace_copy import generate_marketplace_copy
from core.styles import (
    apply_sheet_layout,
    currency_cell_style,
    instruction_box_style,
    kpi_card_style,
    label_style,
    percentage_cell_style,
    placeholder_style,
    section_header_style,
    table_header_style,
    title_style,
    value_style,
    warning_cell_style,
)

OUTPUTS_DIR = Path(__file__).resolve().parents[1] / "outputs"
BOOKKEEPING = "Small Business Bookkeeping / Profit & Loss Dashboard"
SALES = "Sales Dashboard"
INVENTORY = "Inventory Management Dashboard"
INVOICE = "Invoice & Receipt Template with Logo"
CRM = "CRM Customer Tracker"


def build_workbook(template_type, business_details, palette_name, style_name, uploaded_logo=None):
    """Build the selected sellable workbook template."""
    palette = PALETTES[palette_name]
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    if template_type == BOOKKEEPING:
        _build_bookkeeping(workbook, template_type, business_details, palette, style_name, uploaded_logo)
    elif template_type == SALES:
        _build_sales(workbook, template_type, business_details, palette, style_name, uploaded_logo)
    elif template_type == INVENTORY:
        _build_inventory(workbook, template_type, business_details, palette, style_name, uploaded_logo)
    elif template_type == INVOICE:
        _build_invoice(workbook, template_type, business_details, palette, style_name, uploaded_logo)
    elif template_type == CRM:
        _build_crm(workbook, template_type, business_details, palette, style_name, uploaded_logo)
    else:
        raise ValueError(f"Unsupported template type: {template_type}")

    filename = _make_filename(template_type, business_details.get("business_name"))
    output_bytes = _workbook_to_bytes(workbook)
    saved_path = save_workbook(output_bytes, filename)
    generate_marketplace_copy(template_type)
    return filename, output_bytes, saved_path


def create_sheet(workbook, title, widths=None, tab_color=None):
    worksheet = workbook.create_sheet(title=title)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A4"
    if widths:
        apply_sheet_layout(worksheet, widths)
    if tab_color:
        worksheet.sheet_properties.tabColor = tab_color
    return worksheet


def add_title(worksheet, title, palette, cell_range="A1:G1"):
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    worksheet[start_cell] = title
    title_style(worksheet[start_cell], palette)
    worksheet.row_dimensions[worksheet[start_cell].row].height = 34


def add_business_header(worksheet, business_details, template_type, style_name, palette, uploaded_logo=None):
    worksheet.merge_cells("B3:E3")
    worksheet["B3"] = template_type
    section_header_style(worksheet["B3"], palette)
    details = [
        ("Business Name", business_details.get("business_name", "")),
        ("Phone", business_details.get("phone", "")),
        ("Email", business_details.get("email", "")),
        ("Address", business_details.get("address", "")),
        ("Website", business_details.get("website", "")),
        ("Currency", business_details.get("currency", "")),
        ("Tax/VAT/GST %", business_details.get("tax_rate", 0)),
        ("Style", style_name),
    ]
    for row, (label, value) in enumerate(details, start=4):
        worksheet[f"B{row}"] = label
        worksheet[f"C{row}"] = value
        label_style(worksheet[f"B{row}"], palette)
        value_style(worksheet[f"C{row}"], palette)
    worksheet["C10"] = float(business_details.get("tax_rate", 0) or 0) / 100
    percentage_cell_style(worksheet["C10"], palette)
    add_logo_or_placeholder(worksheet, uploaded_logo, palette, anchor="F3", placeholder_range="F3:G6")


def add_kpi_card(worksheet, cell_range, label, value, palette, currency_code="USD", value_type="currency"):
    start_cell, end_cell = cell_range.split(":")
    start_col = worksheet[start_cell].column
    start_row = worksheet[start_cell].row
    end_col = worksheet[end_cell].column
    end_row = worksheet[end_cell].row
    worksheet.merge_cells(f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}")
    worksheet.merge_cells(f"{get_column_letter(start_col)}{start_row + 1}:{get_column_letter(end_col)}{end_row}")
    label_cell = worksheet.cell(start_row, start_col, label)
    value_cell = worksheet.cell(start_row + 1, start_col, value)
    kpi_card_style(label_cell, palette, is_value=False)
    kpi_card_style(value_cell, palette, is_value=True)
    if value_type == "currency":
        currency_cell_style(value_cell, palette, currency_code)
    elif value_type == "percent":
        percentage_cell_style(value_cell, palette)
    else:
        value_style(value_cell, palette)
    value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(worksheet, start_row, start_col, headers, rows, palette, currency_columns=None, percentage_columns=None):
    currency_columns = set(currency_columns or [])
    percentage_columns = set(percentage_columns or [])
    for offset, header in enumerate(headers):
        cell = worksheet.cell(row=start_row, column=start_col + offset, value=header)
        table_header_style(cell, palette)
    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            cell = worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
            header = headers[col_offset]
            if header in currency_columns:
                currency_cell_style(cell, palette)
            elif header in percentage_columns:
                percentage_cell_style(cell, palette)
            else:
                value_style(cell, palette)
    return start_row + len(rows) + 2


def add_chart_placeholder(worksheet, cell_range, title, palette):
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    worksheet[start_cell] = title
    placeholder_style(worksheet[start_cell], palette)


def add_instruction_text(worksheet, cell_range, text, palette, warning=False):
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    worksheet[start_cell] = text
    if warning:
        warning_cell_style(worksheet[start_cell], palette)
    else:
        instruction_box_style(worksheet[start_cell], palette)


def add_logo_or_placeholder_to_sheet(worksheet, uploaded_logo, palette, anchor="F2", placeholder_range="F2:G5"):
    return add_logo_or_placeholder(worksheet, uploaded_logo, palette, anchor, placeholder_range)


def save_workbook(workbook_or_bytes, filename, outputs_dir=OUTPUTS_DIR):
    outputs_dir.mkdir(exist_ok=True)
    file_path = outputs_dir / filename
    if isinstance(workbook_or_bytes, bytes):
        file_path.write_bytes(workbook_or_bytes)
    else:
        workbook_or_bytes.save(file_path)
    return file_path


def _dashboard(workbook, template_type, details, palette, style_name, logo):
    sheet = create_sheet(workbook, "Dashboard", {"A": 4, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18}, palette["primary"])
    add_title(sheet, "Excel Dashboard Studio", palette, "A1:G1")
    add_business_header(sheet, details, template_type, style_name, palette, logo)
    sheet.print_area = "A1:G38"
    return sheet


def _instructions(workbook, template_type, palette, guidance):
    sheet = create_sheet(workbook, "Instructions", {"A": 6, "B": 32, "C": 88}, palette["primary"])
    add_title(sheet, "Instructions", palette, "A1:C1")
    items = [
        ("How to enter data", "Replace sample rows in the input sheets. Add new rows below the examples and copy formulas down when needed."),
        ("Which sheets to update", "Use only the input and lookup sheets listed below. Dashboard and summary sheets are formula-driven."),
        ("Lookup formulas", "IDs connect sheets with XLOOKUP, SUMIF, SUMIFS, COUNTIF, and COUNTIFS formulas. Keep ID values consistent."),
        ("Customize logo, colors, currency, and tax", "Generate a new workbook from Streamlit with your selected logo, palette, currency, and Tax/VAT/GST percentage."),
        ("Print or export dashboard", "Use the preset print area on Dashboard, Invoice, or Receipt, then print or export to PDF from Excel."),
        ("Avoid breaking formulas", "Do not rename linked sheets, delete headers, cut formula cells, or change ID formats without updating linked formulas."),
        ("Selected template", template_type),
    ]
    row = 3
    for title, body in items:
        sheet[f"B{row}"] = title
        sheet[f"C{row}"] = body
        section_header_style(sheet[f"B{row}"], palette)
        instruction_box_style(sheet[f"C{row}"], palette)
        sheet.row_dimensions[row].height = 45
        row += 1
    row += 1
    for name, text in guidance:
        sheet[f"B{row}"] = name
        sheet[f"C{row}"] = text
        label_style(sheet[f"B{row}"], palette)
        value_style(sheet[f"C{row}"], palette)
        row += 1
    add_instruction_text(sheet, f"B{row + 1}:C{row + 2}", "Formula safety: keep lookup keys stable and edit input rows instead of overwriting formulas.", palette, warning=True)
    return sheet


def _style_money(sheet, columns, start, end, palette, currency):
    for column in columns:
        for row in range(start, end + 1):
            currency_cell_style(sheet[f"{column}{row}"], palette, currency)


def _style_percent(sheet, columns, start, end, palette):
    for column in columns:
        for row in range(start, end + 1):
            percentage_cell_style(sheet[f"{column}{row}"], palette)


def _validation(sheet, cell_range, formula_or_csv):
    formula = formula_or_csv if formula_or_csv.startswith('"') else f"={formula_or_csv}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _line_chart(sheet, title, min_col, max_col, min_row, max_row, anchor):
    chart = LineChart()
    chart.title = title
    chart.height = 7
    chart.width = 13
    data = Reference(sheet, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(sheet, min_col=1, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, anchor)


def _bar_chart(sheet, title, min_col, max_col, min_row, max_row, anchor):
    chart = BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 13
    data = Reference(sheet, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(sheet, min_col=1, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, anchor)


def _tax(details):
    return float(details.get("tax_rate", 0) or 0) / 100


def _build_bookkeeping(workbook, template_type, details, palette, style_name, logo):
    currency = details.get("currency", "USD")
    tax_rate = _tax(details)
    dash = _dashboard(workbook, template_type, details, palette, style_name, logo)
    income = create_sheet(workbook, "Income", {"A": 14, "B": 16, "C": 16, "D": 24, "E": 28, "F": 16, "G": 14}, palette["secondary"])
    expenses = create_sheet(workbook, "Expenses", {"A": 14, "B": 16, "C": 16, "D": 24, "E": 28, "F": 16, "G": 14}, palette["secondary"])
    categories = create_sheet(workbook, "Categories", {"A": 16, "B": 24, "C": 16, "D": 14}, palette["accent"])
    monthly = create_sheet(workbook, "Monthly_Summary", {"A": 14, "B": 16, "C": 16, "D": 16, "E": 16}, palette["primary"])
    tax = create_sheet(workbook, "Tax_Summary", {"A": 16, "B": 18, "C": 16, "D": 18}, palette["primary"])

    add_title(categories, "Categories", palette, "A1:D1")
    add_table(categories, 3, 1, ["Category_ID", "Category Name", "Type", "Taxable"], [
        ["INC-001", "Product Sales", "Income", "Yes"], ["INC-002", "Service Revenue", "Income", "Yes"], ["INC-003", "Other Income", "Income", "No"],
        ["EXP-001", "Advertising", "Expense", "Yes"], ["EXP-002", "Office Supplies", "Expense", "Yes"], ["EXP-003", "Rent", "Expense", "No"], ["EXP-004", "Software", "Expense", "Yes"],
    ], palette)

    add_title(income, "Income", palette, "A1:G1")
    income_rows = [
        ["2026-01-05", "IN-1001", "INC-001", '=XLOOKUP(C4,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Online product sale", 2400, tax_rate],
        ["2026-01-19", "IN-1002", "INC-002", '=XLOOKUP(C5,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Consulting project", 1850, tax_rate],
        ["2026-02-09", "IN-1003", "INC-001", '=XLOOKUP(C6,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Retail sales", 3150, tax_rate],
        ["2026-03-14", "IN-1004", "INC-002", '=XLOOKUP(C7,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Monthly retainer", 2100, tax_rate],
        ["2026-04-22", "IN-1005", "INC-003", '=XLOOKUP(C8,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Refund credit", 350, 0],
    ]
    add_table(income, 3, 1, ["Date", "Income_ID", "Category_ID", "Category", "Description", "Amount", "Tax Rate"], income_rows, palette, {"Amount"}, {"Tax Rate"})
    _style_money(income, ["F"], 4, 200, palette, currency)
    _style_percent(income, ["G"], 4, 200, palette)
    _validation(income, "C4:C200", "Categories!$A$4:$A$20")

    add_title(expenses, "Expenses", palette, "A1:G1")
    expense_rows = [
        ["2026-01-08", "EX-2001", "EXP-001", '=XLOOKUP(C4,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Ad Network", 620, tax_rate],
        ["2026-01-22", "EX-2002", "EXP-003", '=XLOOKUP(C5,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Main Street Office", 900, 0],
        ["2026-02-03", "EX-2003", "EXP-004", '=XLOOKUP(C6,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Cloud Apps", 180, tax_rate],
        ["2026-03-11", "EX-2004", "EXP-002", '=XLOOKUP(C7,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Office Store", 240, tax_rate],
        ["2026-04-13", "EX-2005", "EXP-001", '=XLOOKUP(C8,Categories!$A$4:$A$20,Categories!$B$4:$B$20,"")', "Campaign Boost", 760, tax_rate],
    ]
    add_table(expenses, 3, 1, ["Date", "Expense_ID", "Category_ID", "Category", "Vendor", "Amount", "Tax Rate"], expense_rows, palette, {"Amount"}, {"Tax Rate"})
    _style_money(expenses, ["F"], 4, 200, palette, currency)
    _style_percent(expenses, ["G"], 4, 200, palette)
    _validation(expenses, "C4:C200", "Categories!$A$4:$A$20")

    add_title(monthly, "Monthly Summary", palette, "A1:E1")
    rows = []
    for i, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1):
        r = i + 3
        rows.append([month, f'=SUMIFS(Income!$F$4:$F$200,Income!$A$4:$A$200,">="&DATE(2026,{i},1),Income!$A$4:$A$200,"<"&EDATE(DATE(2026,{i},1),1))', f'=SUMIFS(Expenses!$F$4:$F$200,Expenses!$A$4:$A$200,">="&DATE(2026,{i},1),Expenses!$A$4:$A$200,"<"&EDATE(DATE(2026,{i},1),1))', f"=B{r}-C{r}", f'=IF(B{r}=0,0,D{r}/B{r})'])
    add_table(monthly, 3, 1, ["Month", "Income", "Expenses", "Net Profit", "Profit Margin"], rows, palette, {"Income", "Expenses", "Net Profit"}, {"Profit Margin"})
    _style_money(monthly, ["B", "C", "D"], 4, 15, palette, currency)
    _style_percent(monthly, ["E"], 4, 15, palette)
    _line_chart(monthly, "Income vs Expenses by Month", 2, 3, 3, 15, "G3")
    _line_chart(monthly, "Profit Trend", 4, 4, 3, 15, "G18")

    add_title(tax, "Tax Summary", palette, "A1:D1")
    add_table(tax, 3, 1, ["Type", "Taxable Total", "Tax Rate Avg", "Estimated Tax"], [["Income", "=SUM(Income!F4:F200)", "=AVERAGE(Income!G4:G200)", "=B4*C4"], ["Expenses", "=SUM(Expenses!F4:F200)", "=AVERAGE(Expenses!G4:G200)", "=B5*C5"], ["Net Estimate", "=B4-B5", "", "=D4-D5"]], palette, {"Taxable Total", "Estimated Tax"}, {"Tax Rate Avg"})
    _style_money(tax, ["B", "D"], 4, 6, palette, currency)
    _style_percent(tax, ["C"], 4, 6, palette)

    for cell_range, label, value, value_type in [
        ("B13:C15", "Total Income", "=SUM(Income!F4:F200)", "currency"), ("D13:E15", "Total Expenses", "=SUM(Expenses!F4:F200)", "currency"), ("F13:G15", "Net Profit", "=B9-D9", "currency"),
        ("B17:C19", "Profit Margin", "=IF(B9=0,0,F9/B9)", "percent"), ("D17:E19", "Monthly Income", "=SUM(Monthly_Summary!B4:B15)", "currency"), ("F17:G19", "Monthly Expenses", "=SUM(Monthly_Summary!C4:C15)", "currency"),
        ("B21:C23", "Top Expense Category", '=INDEX(Expenses!D4:D200,MATCH(MAX(Expenses!F4:F200),Expenses!F4:F200,0))', "text"), ("D21:E23", "Tax/VAT/GST Estimate", "=Tax_Summary!D6", "currency"),
    ]:
        add_kpi_card(dash, cell_range, label, value, palette, currency, value_type)
    add_chart_placeholder(dash, "B26:D34", "Income vs Expenses by Month", palette)
    add_chart_placeholder(dash, "E26:G34", "Expense Breakdown by Category", palette)
    add_instruction_text(dash, "B36:G38", "Cash flow summary and chart-ready data live in Monthly_Summary and Tax_Summary.", palette)
    _instructions(workbook, template_type, palette, [("Income", "Update income transactions."), ("Expenses", "Update expense transactions."), ("Categories", "Maintain Category_ID lookup values."), ("Monthly_Summary", "Review monthly profit and cash flow formulas."), ("Tax_Summary", "Review tax estimate formulas.")])


def _build_sales(workbook, template_type, details, palette, style_name, logo):
    currency = details.get("currency", "USD")
    dash = _dashboard(workbook, template_type, details, palette, style_name, logo)
    sales = create_sheet(workbook, "Sales_Data", {"A": 14, "B": 14, "C": 14, "D": 22, "E": 14, "F": 22, "G": 16, "H": 14, "I": 14, "J": 14, "K": 14, "L": 16}, palette["secondary"])
    customers = create_sheet(workbook, "Customers", {"A": 14, "B": 24, "C": 18, "D": 18, "E": 24}, palette["accent"])
    products = create_sheet(workbook, "Products", {"A": 14, "B": 24, "C": 18, "D": 14, "E": 14}, palette["accent"])
    monthly = create_sheet(workbook, "Monthly_Summary", {"A": 14, "B": 16, "C": 16, "D": 16, "E": 18}, palette["primary"])

    add_title(customers, "Customers", palette, "A1:E1")
    add_table(customers, 3, 1, ["Customer_ID", "Customer Name", "Region", "Segment", "Email"], [["CUST-001", "Northstar Retail", "North", "Retail", "buyer@northstar.test"], ["CUST-002", "Urban Services", "West", "Service", "ops@urban.test"], ["CUST-003", "Metro Foods", "East", "Wholesale", "ap@metro.test"], ["CUST-004", "Bright Office", "South", "Business", "team@bright.test"]], palette)
    add_title(products, "Products", palette, "A1:E1")
    add_table(products, 3, 1, ["Product_ID", "Product Name", "Category", "Unit Price", "Unit Cost"], [["PROD-001", "Starter Dashboard", "Templates", 79, 18], ["PROD-002", "Premium Dashboard", "Templates", 149, 36], ["PROD-003", "Setup Service", "Services", 300, 95], ["PROD-004", "Training Session", "Services", 180, 55]], palette, {"Unit Price", "Unit Cost"})
    _style_money(products, ["D", "E"], 4, 50, palette, currency)

    add_title(sales, "Sales Data", palette, "A1:L1")
    rows = []
    for idx, data in enumerate([("2026-01-07", "SO-1001", "CUST-001", "PROD-001", 6), ("2026-01-18", "SO-1002", "CUST-002", "PROD-003", 2), ("2026-02-09", "SO-1003", "CUST-003", "PROD-002", 4), ("2026-03-14", "SO-1004", "CUST-004", "PROD-004", 3)], start=4):
        date, order_id, customer_id, product_id, qty = data
        rows.append([date, order_id, customer_id, f'=XLOOKUP(C{idx},Customers!$A$4:$A$50,Customers!$B$4:$B$50,"")', product_id, f'=XLOOKUP(E{idx},Products!$A$4:$A$50,Products!$B$4:$B$50,"")', f'=XLOOKUP(E{idx},Products!$A$4:$A$50,Products!$C$4:$C$50,"")', f'=XLOOKUP(C{idx},Customers!$A$4:$A$50,Customers!$C$4:$C$50,"")', qty, f'=XLOOKUP(E{idx},Products!$A$4:$A$50,Products!$D$4:$D$50,0)', f'=XLOOKUP(E{idx},Products!$A$4:$A$50,Products!$E$4:$E$50,0)', f"=(J{idx}-K{idx})*I{idx}"])
    add_table(sales, 3, 1, ["Order Date", "Order_ID", "Customer_ID", "Customer Name", "Product_ID", "Product Name", "Category", "Region", "Quantity", "Unit Price", "Unit Cost", "Profit"], rows, palette, {"Unit Price", "Unit Cost", "Profit"})
    _style_money(sales, ["J", "K", "L"], 4, 200, palette, currency)
    _validation(sales, "C4:C200", "Customers!$A$4:$A$50")
    _validation(sales, "E4:E200", "Products!$A$4:$A$50")

    add_title(monthly, "Monthly Summary", palette, "A1:E1")
    rows = []
    for i, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1):
        r = i + 3
        rows.append([month, f'=SUMPRODUCT((MONTH(Sales_Data!$A$4:$A$200)={i})*(Sales_Data!$I$4:$I$200*Sales_Data!$J$4:$J$200))', f'=COUNTIFS(Sales_Data!$A$4:$A$200,">="&DATE(2026,{i},1),Sales_Data!$A$4:$A$200,"<"&EDATE(DATE(2026,{i},1),1))', f'=SUMIFS(Sales_Data!$L$4:$L$200,Sales_Data!$A$4:$A$200,">="&DATE(2026,{i},1),Sales_Data!$A$4:$A$200,"<"&EDATE(DATE(2026,{i},1),1))', f"=IF(C{r}=0,0,B{r}/C{r})"])
    add_table(monthly, 3, 1, ["Month", "Sales", "Orders", "Profit", "Average Order Value"], rows, palette, {"Sales", "Profit", "Average Order Value"})
    _style_money(monthly, ["B", "D", "E"], 4, 15, palette, currency)
    _line_chart(monthly, "Monthly Sales Trend", 2, 2, 3, 15, "G3")

    for cell_range, label, value, value_type in [("B13:C15", "Total Sales", "=SUMPRODUCT(Sales_Data!I4:I200,Sales_Data!J4:J200)", "currency"), ("D13:E15", "Total Orders", "=COUNTA(Sales_Data!B4:B200)", "number"), ("F13:G15", "Average Order Value", "=IF(D9=0,0,B9/D9)", "currency"), ("B17:C19", "Total Profit", "=SUM(Sales_Data!L4:L200)", "currency"), ("D17:E19", "Profit Margin", "=IF(B9=0,0,B13/B9)", "percent"), ("F17:G19", "Top Product", '=INDEX(Sales_Data!F4:F200,MATCH(MAX(Sales_Data!L4:L200),Sales_Data!L4:L200,0))', "text"), ("B21:C23", "Top Customer", '=INDEX(Sales_Data!D4:D200,MATCH(MAX(Sales_Data!L4:L200),Sales_Data!L4:L200,0))', "text"), ("D21:E23", "Best Month", '=INDEX(Monthly_Summary!A4:A15,MATCH(MAX(Monthly_Summary!B4:B15),Monthly_Summary!B4:B15,0))', "text")]:
        add_kpi_card(dash, cell_range, label, value, palette, currency, value_type)
    add_chart_placeholder(dash, "B26:D34", "Sales by Product", palette)
    add_chart_placeholder(dash, "E26:G34", "Sales by Customer / Region", palette)
    add_instruction_text(dash, "B36:G38", "Sales_Data links Customers and Products with XLOOKUP. Monthly_Summary provides chart-ready trend data.", palette)
    _instructions(workbook, template_type, palette, [("Sales_Data", "Enter order rows and use Customer_ID/Product_ID dropdowns."), ("Customers", "Maintain customer lookup details."), ("Products", "Maintain product price and cost lookup details."), ("Monthly_Summary", "Review monthly sales, profit, and average order value.")])


def _build_inventory(workbook, template_type, details, palette, style_name, logo):
    currency = details.get("currency", "USD")
    dash = _dashboard(workbook, template_type, details, palette, style_name, logo)
    products = create_sheet(workbook, "Products", {"A": 14, "B": 24, "C": 18, "D": 14, "E": 22, "F": 14, "G": 14, "H": 16}, palette["secondary"])
    stock_in = create_sheet(workbook, "Stock_In", {"A": 14, "B": 14, "C": 24, "D": 14, "E": 14, "F": 24}, palette["accent"])
    stock_out = create_sheet(workbook, "Stock_Out", {"A": 14, "B": 14, "C": 24, "D": 14, "E": 14, "F": 24}, palette["accent"])
    suppliers = create_sheet(workbook, "Suppliers", {"A": 14, "B": 24, "C": 24, "D": 18}, palette["secondary"])
    reorder = create_sheet(workbook, "Reorder_Report", {"A": 14, "B": 24, "C": 14, "D": 14, "E": 18, "F": 16}, palette["primary"])

    add_title(suppliers, "Suppliers", palette, "A1:D1")
    add_table(suppliers, 3, 1, ["Supplier_ID", "Supplier Name", "Contact", "Region"], [["SUP-001", "Acme Supply", "sales@acme.test", "North"], ["SUP-002", "Global Goods", "team@global.test", "West"], ["SUP-003", "Local Wholesale", "orders@local.test", "South"]], palette)
    add_title(products, "Products", palette, "A1:H1")
    product_rows = [["INV-001", "Laptop Stand", "Accessories", "SUP-001", '=XLOOKUP(D4,Suppliers!$A$4:$A$50,Suppliers!$B$4:$B$50,"")', 18, 25, '=SUMIF(Stock_In!$B$4:$B$200,A4,Stock_In!$D$4:$D$200)-SUMIF(Stock_Out!$B$4:$B$200,A4,Stock_Out!$D$4:$D$200)'], ["INV-002", "Desk Lamp", "Office", "SUP-002", '=XLOOKUP(D5,Suppliers!$A$4:$A$50,Suppliers!$B$4:$B$50,"")', 24, 15, '=SUMIF(Stock_In!$B$4:$B$200,A5,Stock_In!$D$4:$D$200)-SUMIF(Stock_Out!$B$4:$B$200,A5,Stock_Out!$D$4:$D$200)'], ["INV-003", "Notebook Pack", "Stationery", "SUP-003", '=XLOOKUP(D6,Suppliers!$A$4:$A$50,Suppliers!$B$4:$B$50,"")', 4, 80, '=SUMIF(Stock_In!$B$4:$B$200,A6,Stock_In!$D$4:$D$200)-SUMIF(Stock_Out!$B$4:$B$200,A6,Stock_Out!$D$4:$D$200)'], ["INV-004", "Wireless Mouse", "Accessories", "SUP-001", '=XLOOKUP(D7,Suppliers!$A$4:$A$50,Suppliers!$B$4:$B$50,"")', 12, 30, '=SUMIF(Stock_In!$B$4:$B$200,A7,Stock_In!$D$4:$D$200)-SUMIF(Stock_Out!$B$4:$B$200,A7,Stock_Out!$D$4:$D$200)']]
    add_table(products, 3, 1, ["Product_ID", "Product Name", "Category", "Supplier_ID", "Supplier", "Unit Cost", "Reorder Level", "Current Stock"], product_rows, palette, {"Unit Cost"})
    _style_money(products, ["F"], 4, 100, palette, currency)
    _validation(products, "D4:D100", "Suppliers!$A$4:$A$50")

    for sheet, title, qtys in [(stock_in, "Stock In", [90, 42, 160, 60]), (stock_out, "Stock Out", [22, 12, 95, 40])]:
        add_title(sheet, title, palette, "A1:F1")
        rows = []
        for idx, (pid, qty) in enumerate(zip(["INV-001", "INV-002", "INV-003", "INV-004"], qtys), start=4):
            rows.append([f"2026-0{min(idx-3,3)}-{idx+2:02d}", pid, f'=XLOOKUP(B{idx},Products!$A$4:$A$100,Products!$B$4:$B$100,"")', qty, f'=XLOOKUP(B{idx},Products!$A$4:$A$100,Products!$F$4:$F$100,0)', "Sample movement"])
        add_table(sheet, 3, 1, ["Date", "Product_ID", "Product Name", "Quantity", "Unit Cost", "Notes"], rows, palette, {"Unit Cost"})
        _style_money(sheet, ["E"], 4, 200, palette, currency)
        _validation(sheet, "B4:B200", "Products!$A$4:$A$100")

    add_title(reorder, "Reorder Report", palette, "A1:F1")
    rows = [[f"=Products!A{r}", f"=Products!B{r}", f"=Products!H{r}", f"=Products!G{r}", f'=IF(C{r}=0,"Out of Stock",IF(C{r}<=D{r},"Reorder","OK"))', f"=C{r}*Products!F{r}"] for r in range(4, 8)]
    add_table(reorder, 3, 1, ["Product_ID", "Product Name", "Current Stock", "Reorder Level", "Status", "Stock Value"], rows, palette, {"Stock Value"})
    _style_money(reorder, ["F"], 4, 100, palette, currency)
    reorder.conditional_formatting.add("E4:E100", FormulaRule(formula=['E4="Reorder"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    reorder.conditional_formatting.add("E4:E100", FormulaRule(formula=['E4="Out of Stock"'], fill=PatternFill("solid", fgColor="FF9999")))

    for cell_range, label, value, value_type in [("B13:C15", "Total Products", "=COUNTA(Products!A4:A100)", "number"), ("D13:E15", "Current Stock Value", "=SUM(Reorder_Report!F4:F100)", "currency"), ("F13:G15", "Low Stock Items", '=COUNTIF(Reorder_Report!E4:E100,"Reorder")', "number"), ("B17:C19", "Out of Stock Items", '=COUNTIF(Reorder_Report!E4:E100,"Out of Stock")', "number"), ("D17:E19", "Fast-moving Products", '=INDEX(Products!B4:B100,MATCH(MAX(Stock_Out!D4:D200),Stock_Out!D4:D200,0))', "text"), ("F17:G19", "Dead Stock Products", '=INDEX(Products!B4:B100,MATCH(MIN(Stock_Out!D4:D200),Stock_Out!D4:D200,0))', "text"), ("B21:C23", "Reorder Required", '=COUNTIF(Reorder_Report!E4:E100,"Reorder")', "number")]:
        add_kpi_card(dash, cell_range, label, value, palette, currency, value_type)
    add_chart_placeholder(dash, "B26:D34", "Stock Level by Product / Category", palette)
    add_chart_placeholder(dash, "E26:G34", "Supplier-wise Stock Value", palette)
    add_instruction_text(dash, "B36:G38", "Stock_In and Stock_Out link to Products by Product_ID. Products links to Suppliers by Supplier_ID.", palette)
    _instructions(workbook, template_type, palette, [("Products", "Maintain product, supplier, cost, and reorder levels."), ("Stock_In", "Record purchases and stock increases."), ("Stock_Out", "Record sales, usage, and stock decreases."), ("Suppliers", "Maintain supplier lookup details."), ("Reorder_Report", "Review reorder status and stock value formulas.")])


def _build_invoice(workbook, template_type, details, palette, style_name, logo):
    currency = details.get("currency", "USD")
    tax_rate = _tax(details)
    invoice = create_sheet(workbook, "Invoice", {"A": 14, "B": 20, "C": 24, "D": 12, "E": 14, "F": 14, "G": 16}, palette["primary"])
    receipt = create_sheet(workbook, "Receipt", {"A": 12, "B": 22, "C": 22, "D": 18, "E": 18}, palette["secondary"])
    customers = create_sheet(workbook, "Customers", {"A": 14, "B": 24, "C": 28, "D": 18, "E": 28}, palette["accent"])
    products = create_sheet(workbook, "Products", {"A": 14, "B": 24, "C": 16, "D": 16, "E": 16}, palette["accent"])
    log = create_sheet(workbook, "Invoice_Log", {"A": 14, "B": 14, "C": 24, "D": 16, "E": 16, "F": 16, "G": 18}, palette["primary"])
    tracker = create_sheet(workbook, "Payment_Tracker", {"A": 24, "B": 18, "C": 40, "D": 18, "E": 18}, palette["primary"])

    add_title(invoice, "Invoice", palette, "A1:G1")
    add_logo_or_placeholder(invoice, logo, palette, "A3", "A3:B6")
    for row, value in enumerate([details.get("business_name", "Business Name"), details.get("address", "Address"), details.get("phone", "Phone"), details.get("email", "Email"), details.get("website", "Website")], start=3):
        invoice[f"D{row}"] = value
        value_style(invoice[f"D{row}"], palette)
    for row, (label, value) in enumerate([("Invoice #", "INV-1001"), ("Invoice Date", "2026-01-15"), ("Due Date", "2026-01-30"), ("Payment Status", "Unpaid")], start=3):
        invoice[f"F{row}"] = label
        invoice[f"G{row}"] = value
        label_style(invoice[f"F{row}"], palette)
        value_style(invoice[f"G{row}"], palette)
    invoice["A8"] = "Customer_ID"
    invoice["B8"] = "CUST-001"
    invoice["A9"] = "Customer Name"
    invoice["B9"] = '=XLOOKUP(B8,Customers!$A$4:$A$50,Customers!$B$4:$B$50,"")'
    for cell in ["A8", "A9"]:
        label_style(invoice[cell], palette)
    for cell in ["B8", "B9"]:
        value_style(invoice[cell], palette)
    rows = []
    for row, pid in enumerate(["PROD-001", "PROD-002", "PROD-003", ""], start=12):
        rows.append([pid, f'=IF(A{row}="","",XLOOKUP(A{row},Products!$A$4:$A$50,Products!$B$4:$B$50,""))', 1 if pid else "", f'=IF(A{row}="","",XLOOKUP(A{row},Products!$A$4:$A$50,Products!$D$4:$D$50,0))', 0, tax_rate, f'=IF(A{row}="","",C{row}*D{row}-E{row}+(C{row}*D{row}-E{row})*F{row})'])
    add_table(invoice, 11, 1, ["Product_ID", "Item", "Quantity", "Unit Price", "Discount", "Tax/VAT/GST", "Line Total"], rows, palette, {"Unit Price", "Discount", "Line Total"}, {"Tax/VAT/GST"})
    _style_money(invoice, ["D", "E", "G"], 12, 30, palette, currency)
    _style_percent(invoice, ["F"], 12, 30, palette)
    for row, label, formula in [(18, "Subtotal", "=SUMPRODUCT(C12:C15,D12:D15)-SUM(E12:E15)"), (19, "Tax/VAT/GST", "=SUMPRODUCT((C12:C15*D12:D15-E12:E15),F12:F15)"), (20, "Grand Total", "=SUM(G12:G15)")]:
        invoice[f"F{row}"] = label
        invoice[f"G{row}"] = formula
        label_style(invoice[f"F{row}"], palette)
        currency_cell_style(invoice[f"G{row}"], palette, currency)
    _validation(invoice, "B8", "Customers!$A$4:$A$50")
    _validation(invoice, "A12:A30", "Products!$A$4:$A$50")
    _validation(invoice, "G6", '"Paid,Unpaid,Partial,Overdue"')
    invoice.print_area = "A1:G22"

    add_title(receipt, "Receipt", palette, "A1:E1")
    add_logo_or_placeholder(receipt, logo, palette, "A3", "A3:B5")
    for row, (label, value) in enumerate([("Receipt #", "RCPT-1001"), ("Date", "2026-01-16"), ("Received From", "=Invoice!B9"), ("Invoice #", "=Invoice!G3"), ("Amount Paid", "=Invoice!G20"), ("Payment Method", "Bank Transfer")], start=7):
        receipt[f"B{row}"] = label
        receipt[f"C{row}"] = value
        label_style(receipt[f"B{row}"], palette)
        value_style(receipt[f"C{row}"], palette)
    currency_cell_style(receipt["C11"], palette, currency)
    add_instruction_text(receipt, "B15:E17", "Printable receipt area. Export this sheet to PDF after payment is received.", palette)
    receipt.print_area = "A1:E18"

    add_title(customers, "Customers", palette, "A1:E1")
    add_table(customers, 3, 1, ["Customer_ID", "Customer Name", "Billing Address", "Phone", "Email"], [["CUST-001", "Northstar Retail", "10 Market Road", "555-1001", "billing@northstar.test"], ["CUST-002", "Urban Services", "44 City Avenue", "555-1002", "pay@urban.test"]], palette)
    add_title(products, "Products", palette, "A1:E1")
    add_table(products, 3, 1, ["Product_ID", "Item Name", "Category", "Unit Price", "Taxable"], [["PROD-001", "Dashboard Template", "Digital", 129, "Yes"], ["PROD-002", "Customization", "Service", 250, "Yes"], ["PROD-003", "Support Package", "Service", 99, "No"]], palette, {"Unit Price"})
    _style_money(products, ["D"], 4, 100, palette, currency)
    add_title(log, "Invoice Log", palette, "A1:G1")
    add_table(log, 3, 1, ["Invoice_ID", "Customer_ID", "Customer Name", "Invoice Date", "Due Date", "Amount", "Payment Status"], [["INV-1001", "CUST-001", '=XLOOKUP(B4,Customers!$A$4:$A$50,Customers!$B$4:$B$50,"")', "2026-01-15", "2026-01-30", "=Invoice!G20", "Unpaid"], ["INV-1000", "CUST-002", '=XLOOKUP(B5,Customers!$A$4:$A$50,Customers!$B$4:$B$50,"")', "2026-01-05", "2026-01-20", 349, "Paid"]], palette, {"Amount"})
    _style_money(log, ["F"], 4, 200, palette, currency)
    _validation(log, "B4:B200", "Customers!$A$4:$A$50")
    _validation(log, "G4:G200", '"Paid,Unpaid,Partial,Overdue"')
    log.conditional_formatting.add("G4:G200", FormulaRule(formula=['G4="Overdue"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    add_title(tracker, "Payment Tracker", palette, "A1:E1")
    add_table(tracker, 3, 1, ["Metric", "Value", "Notes", "Status", "Balance"], [["Total invoices", "=COUNTA(Invoice_Log!A4:A200)", "All logged invoices", "", ""], ["Paid invoices", '=COUNTIF(Invoice_Log!G4:G200,"Paid")', "Paid count", "", ""], ["Unpaid invoices", '=COUNTIF(Invoice_Log!G4:G200,"Unpaid")', "Unpaid count", "", ""], ["Total invoice value", "=SUM(Invoice_Log!F4:F200)", "Total billed", "", ""], ["Outstanding balance", '=SUMIF(Invoice_Log!G4:G200,"<>Paid",Invoice_Log!F4:F200)', "Needs follow-up", "", ""]], palette, {"Value", "Balance"})
    _style_money(tracker, ["B"], 7, 8, palette, currency)
    _instructions(workbook, template_type, palette, [("Invoice", "Select Customer_ID and Product_ID values, then print or export."), ("Receipt", "Use after payment as printable proof of payment."), ("Customers", "Maintain customer lookup details."), ("Products", "Maintain price lookup details."), ("Invoice_Log", "Track invoice values and payment status."), ("Payment_Tracker", "Review invoice KPIs and outstanding balances.")])


def _build_crm(workbook, template_type, details, palette, style_name, logo):
    currency = details.get("currency", "USD")
    dash = _dashboard(workbook, template_type, details, palette, style_name, logo)
    customers = create_sheet(workbook, "Customers", {"A": 14, "B": 24, "C": 18, "D": 18, "E": 18}, palette["secondary"])
    leads = create_sheet(workbook, "Leads", {"A": 14, "B": 24, "C": 18, "D": 18, "E": 16}, palette["accent"])
    deals = create_sheet(workbook, "Deals", {"A": 14, "B": 14, "C": 24, "D": 18, "E": 16, "F": 16, "G": 16}, palette["secondary"])
    followups = create_sheet(workbook, "Followups", {"A": 14, "B": 14, "C": 24, "D": 16, "E": 18, "F": 30}, palette["accent"])
    pipeline = create_sheet(workbook, "Sales_Pipeline", {"A": 18, "B": 16, "C": 18, "D": 16}, palette["primary"])

    add_title(customers, "Customers", palette, "A1:E1")
    add_table(customers, 3, 1, ["Customer_ID", "Customer Name", "Industry", "Owner", "Status"], [["CRM-001", "Northstar Retail", "Retail", "Ayesha", "Active"], ["CRM-002", "Urban Services", "Services", "Bilal", "Active"], ["CRM-003", "Metro Foods", "Food", "Ayesha", "Prospect"], ["CRM-004", "Bright Office", "Business", "Sara", "Active"]], palette)
    add_title(leads, "Leads", palette, "A1:E1")
    add_table(leads, 3, 1, ["Lead_ID", "Lead Name", "Source", "Status", "Created Date"], [["LEAD-001", "Website Trial", "Website", "New", "2026-01-03"], ["LEAD-002", "Referral Lead", "Referral", "Qualified", "2026-01-10"], ["LEAD-003", "Trade Show", "Event", "New", "2026-02-01"]], palette)
    _validation(leads, "D4:D200", '"New,Qualified,Converted,Lost"')
    add_title(deals, "Deals", palette, "A1:G1")
    deal_rows = [["DEAL-001", "CRM-001", '=XLOOKUP(B4,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "Proposal", "Open", 4200, "2026-02-15"], ["DEAL-002", "CRM-002", '=XLOOKUP(B5,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "Negotiation", "Open", 2800, "2026-03-05"], ["DEAL-003", "CRM-003", '=XLOOKUP(B6,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "Closed Won", "Won", 5100, "2026-01-25"], ["DEAL-004", "CRM-004", '=XLOOKUP(B7,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "Closed Lost", "Lost", 1200, "2026-01-30"]]
    add_table(deals, 3, 1, ["Deal_ID", "Customer_ID", "Customer Name", "Stage", "Status", "Value", "Close Date"], deal_rows, palette, {"Value"})
    _style_money(deals, ["F"], 4, 200, palette, currency)
    _validation(deals, "B4:B200", "Customers!$A$4:$A$100")
    _validation(deals, "D4:D200", '"Prospecting,Proposal,Negotiation,Closed Won,Closed Lost"')
    _validation(deals, "E4:E200", '"Open,Won,Lost"')
    deals.conditional_formatting.add("E4:E200", FormulaRule(formula=['E4="Lost"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    add_title(followups, "Followups", palette, "A1:F1")
    followup_rows = [["FU-001", "CRM-001", '=XLOOKUP(B4,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "2026-02-01", "Pending", "Send proposal update"], ["FU-002", "CRM-002", '=XLOOKUP(B5,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "2026-02-04", "Pending", "Schedule call"], ["FU-003", "CRM-003", '=XLOOKUP(B6,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "2026-01-28", "Done", "Closed deal"], ["FU-004", "CRM-004", '=XLOOKUP(B7,Customers!$A$4:$A$100,Customers!$B$4:$B$100,"")', "2026-02-09", "Pending", "Re-engage later"]]
    add_table(followups, 3, 1, ["Followup_ID", "Customer_ID", "Customer Name", "Due Date", "Status", "Notes"], followup_rows, palette)
    _validation(followups, "B4:B200", "Customers!$A$4:$A$100")
    _validation(followups, "E4:E200", '"Pending,Done,Overdue"')
    followups.conditional_formatting.add("E4:E200", FormulaRule(formula=['E4="Overdue"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    add_title(pipeline, "Sales Pipeline", palette, "A1:D1")
    stages = ["Prospecting", "Proposal", "Negotiation", "Closed Won", "Closed Lost"]
    pipe_rows = [[stage, f'=COUNTIF(Deals!$D$4:$D$200,A{idx})', f'=SUMIF(Deals!$D$4:$D$200,A{idx},Deals!$F$4:$F$200)', f'=IF(SUM(B$4:B$8)=0,0,B{idx}/SUM(B$4:B$8))'] for idx, stage in enumerate(stages, start=4)]
    add_table(pipeline, 3, 1, ["Stage", "Deal Count", "Pipeline Value", "Share"], pipe_rows, palette, {"Pipeline Value"}, {"Share"})
    _style_money(pipeline, ["C"], 4, 20, palette, currency)
    _style_percent(pipeline, ["D"], 4, 20, palette)
    _bar_chart(pipeline, "Pipeline Value by Stage", 3, 3, 3, 8, "F3")

    for cell_range, label, value, value_type in [("B13:C15", "Total Customers", "=COUNTA(Customers!A4:A100)", "number"), ("D13:E15", "New Leads", '=COUNTIF(Leads!D4:D200,"New")', "number"), ("F13:G15", "Active Deals", '=COUNTIF(Deals!E4:E200,"Open")', "number"), ("B17:C19", "Won Deals", '=COUNTIF(Deals!E4:E200,"Won")', "number"), ("D17:E19", "Lost Deals", '=COUNTIF(Deals!E4:E200,"Lost")', "number"), ("F17:G19", "Pending Follow-ups", '=COUNTIF(Followups!E4:E200,"Pending")', "number"), ("B21:C23", "Pipeline Value", '=SUMIF(Deals!E4:E200,"Open",Deals!F4:F200)', "currency"), ("D21:E23", "Conversion Rate", '=IF(COUNTA(Deals!A4:A200)=0,0,COUNTIF(Deals!E4:E200,"Won")/COUNTA(Deals!A4:A200))', "percent")]:
        add_kpi_card(dash, cell_range, label, value, palette, currency, value_type)
    add_chart_placeholder(dash, "B26:D34", "Deals by Status", palette)
    add_chart_placeholder(dash, "E26:G34", "Pipeline Value / Leads / Follow-ups", palette)
    add_instruction_text(dash, "B36:G38", "Deals and Followups link to Customers by Customer_ID. Sales_Pipeline provides chart-ready stage totals.", palette)
    _instructions(workbook, template_type, palette, [("Customers", "Maintain customer lookup details."), ("Leads", "Track lead source and qualification status."), ("Deals", "Track deal stage, status, and value."), ("Followups", "Track pending calls, emails, and meetings."), ("Sales_Pipeline", "Review pipeline totals and chart-ready data.")])


def _workbook_to_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _make_filename(template_type, business_name):
    base_name = business_name or template_type
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", base_name).strip("_").lower() or "dashboard"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{safe_name}_{timestamp}.xlsx"



