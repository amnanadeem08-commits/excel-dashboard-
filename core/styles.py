"""Shared workbook styling helpers."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DEFAULT_FONT = "Calibri"


def apply_sheet_layout(worksheet, widths):
    """Apply predictable column widths to a worksheet."""
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def thin_border(color="D9D9D9"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def medium_border(color="BFBFBF"):
    side = Side(style="medium", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def title_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, size=20, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=palette["primary"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = medium_border(palette["primary"])


def section_header_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=palette["secondary"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border(palette["secondary"])


def kpi_card_style(cell, palette, is_value=False):
    cell.fill = PatternFill("solid", fgColor=palette["accent"])
    cell.font = Font(
        name=DEFAULT_FONT,
        size=16 if is_value else 10,
        bold=is_value,
        color=palette["primary"] if is_value else palette["text"],
    )
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = medium_border(palette["secondary"])


def table_header_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=palette["primary"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def currency_cell_style(cell, palette, currency_code="USD"):
    cell.font = Font(name=DEFAULT_FONT, color=palette["text"])
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border()
    cell.number_format = _currency_format(currency_code)


def percentage_cell_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, color=palette["text"])
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border()
    cell.number_format = "0.00%"


def warning_cell_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, bold=True, color="9C0006")
    cell.fill = PatternFill("solid", fgColor="FFC7CE")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border("9C0006")


def instruction_box_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, color=palette["text"])
    cell.fill = PatternFill("solid", fgColor=palette["accent"])
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = thin_border(palette["secondary"])


def label_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, bold=True, color=palette["text"])
    cell.fill = PatternFill("solid", fgColor=palette["accent"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def value_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, color=palette["text"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def placeholder_style(cell, palette):
    cell.font = Font(name=DEFAULT_FONT, bold=True, italic=True, color=palette["secondary"])
    cell.fill = PatternFill("solid", fgColor=palette["accent"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = medium_border(palette["secondary"])


def _currency_format(currency_code):
    symbols = {
        "USD": '"$"#,##0.00',
        "EUR": '"EUR "#,##0.00',
        "GBP": '"GBP "#,##0.00',
        "PKR": '"PKR "#,##0.00',
        "INR": '"INR "#,##0.00',
        "CAD": '"CAD "#,##0.00',
        "AUD": '"AUD "#,##0.00',
        "AED": '"AED "#,##0.00',
    }
    return symbols.get(currency_code, f'"{currency_code} "#,##0.00')
